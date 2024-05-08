from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

workbook = openpyxl.load_workbook("C:/Users/eddyg/Desktop/FYP/GitHub/fyp/YouTube/email_password.xlsx")
worksheet = workbook.active
 
# Iterate over each row in the worksheet
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    # Extract name and email from each row
    name = row[0].value
    email = row[1].value

    driver = webdriver.Firefox()
    driver.get('https://support.google.com/youtube/contact/trademark_complaint')

    full_legal_name = driver.find_element(By.XPATH, '//*[@id="Fulllegalname"]')
    full_legal_name.send_keys(name)

    time.sleep(2)

    email_add = driver.find_element(By.XPATH, '//*[@id="email_prefill"]')
    email_add.send_keys(email)

    time.sleep(2)

    title = driver.find_element(By.XPATH, '//*[@id="Title"]')
    title.send_keys("Mr")

    time.sleep(2)

    company_name = driver.find_element(By.XPATH, '//*[@id="CompanyName"]')
    company_name.send_keys("First Bank Nigeria")

    time.sleep(2)

    trademark_owner = driver.find_element(By.XPATH, '//*[@id="TrademarkOwnerName"]')
    trademark_owner.send_keys("First Bank Nigeria")

    time.sleep(2)

    relationship = driver.find_element(By.XPATH, '//*[@id="OwnerRelationship"]')
    relationship.send_keys("I am reporting on behalf of my organization or client.")

    time.sleep(2)

    nb_trademarks = driver.find_element(By.XPATH , '/html/body/div[2]/div/section/div/div/article/form/div[11]/div')
    nb_trademarks.click()

    time.sleep(2)

    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id=":2"]'))
    )

    opt_1 = driver.find_element(By.XPATH, '//*[@id=":2"]')
    opt_1.click()

    time.sleep(2)

    WebDriverWait(driver , 5).until(
        EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[13]/div'))
    )
    select_brand = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[13]/div')
    select_brand.click()

    time.sleep(2)

    wordmark_and_logo = driver.find_element(By.XPATH, '//*[@id=":10"]')
    wordmark_and_logo.click()

    time.sleep(2)

    yes = driver.find_element(By.XPATH, '//*[@id="trademark_complaint"]/div[14]/fieldset/div[1]/div/label/div[1]')
    yes.click()

    time.sleep(2)

    juridsiction = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[15]/div')
    juridsiction.click()

    time.sleep(2)

    other_opt = driver.find_element(By.XPATH, '//*[@id=":13"]')
    other_opt.click()

    time.sleep(2)

    select_country = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[17]/div')
    select_country.click()

    time.sleep(2)

    nigeria = driver.find_element(By.XPATH, '//*[@id=":181"]')
    nigeria.click()

    time.sleep(2)

    registration_nb = driver.find_element(By.XPATH, '//*[@id="RegistrationNumber_one"]')
    registration_nb.send_keys("70543")

    time.sleep(2)

    type_infriging_content = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[60]/div')
    type_infriging_content.click()

    time.sleep(2)

    channel_opt = driver.find_element(By.XPATH , '/html/body/div[2]/div/section/div/div/article/form/div[60]/div/ol/li[3]')
    channel_opt.click()

    time.sleep(2)

    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '#channel_url_manual'))
    )

    channel_url = driver.find_element(By.CSS_SELECTOR, '#channel_url_manual')
    channel_url.send_keys("https://www.youtube.com/@folorunshomahmud4737")
    
    time.sleep(2)

    desc = driver.find_element(By.XPATH, '//*[@id="AllegedlyInfringed"]')
    desc.send_keys("This channel stole the logo of the organization without permission and used it in all of their videos. Please consider this report.")

    time.sleep(2)

    aff1 = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[70]/fieldset/div/label/label/span[1]')
    aff1.click()

    time.sleep(2)

    aff2 = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[71]/fieldset/div/label/label/span[1]')
    aff2.click()

    time.sleep(2)

    aff3 = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[72]/fieldset/div/label/label/span[1]')
    aff3.click()

    time.sleep(2)

    signature = driver.find_element(By.XPATH, '//*[@id="Signature"]')
    signature.send_keys(name)

    time.sleep(2)

    submit_btn = driver.find_element(By.XPATH, '/html/body/div[2]/div/section/div/div/article/form/div[79]/div/button')
    submit_btn.click()

    time.sleep(3)
                
    driver.quit()

    time.sleep(150)




