from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import pyautogui
import openpyxl
import tiktok_verification

#with verification and many accounts

workbook = openpyxl.load_workbook("C:/Users/eddyg/Desktop/FYP/GitHub/fyp/TikTok/email_password.xlsx")
worksheet = workbook.active
 
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    # Extract name and email from each row
    name = row[0].value
    email = row[1].value
    password = row[2].value

    driver = webdriver.Firefox()
    driver.get('https://www.tiktok.com/legal/report/Copyright')


    emailTxt = driver.find_element(By.TAG_NAME, 'input')
    time.sleep(2)
    emailTxt.send_keys(email)


    time.sleep(5)

    next_btn = driver.find_element(By.CSS_SELECTOR , 'body > div.base-layout-container > main > article > div > form > button')
    next_btn.click()

    #time.sleep(20)

    #url = tiktok_verification.get_outlook_otp(email,password)
    #driver.execute_script("window.open('about:blank', '_blank');")
    #new_tab = driver.window_handles[-1]
    #driver.switch_to.window(new_tab)
    #driver.get(url)  # Replace this URL with the one you want
#
    ## Wait for 10 seconds
    #time.sleep(10)
#
    ## Close the current tab
    #driver.close()
#
    ## Switch back to the original tab
    #driver.switch_to.window(driver.window_handles[0])
#
    #time.sleep(10)
    #done_btn = driver.find_element(By.CSS_SELECTOR, 'button.submit-button:nth-child(5)')
    #done_btn.click()
    #WebDriverWait(driver , 5).until(
    #    EC.presence_of_element_located((By.CLASS_NAME, 'css-1mu1jn5'))
    #)
#
    time.sleep(10)
    problem_selection = driver.find_element(By.CLASS_NAME , 'css-1mu1jn5')
    problem_selection.click()

    time.sleep(5)

    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located ((By.ID , 'tux-2_list_1'))
    )


    subproblem_selection = driver.find_element(By.ID, 'tux-2_list_1')
    subproblem_selection.click()

    time.sleep(5)

    affected = driver.find_element(By.XPATH, '//*[@id="tux-4_button"]')
    affected.click()

    time.sleep(4)
    WebDriverWait(driver , 5).until(
        EC.presence_of_element_located((By.ID, 'tux-4_list_4'))
    )

    affected_select = driver.find_element(By.ID, 'tux-4_list_4')
    affected_select.click()

    time.sleep(4)


    full_name = driver.find_element(By.XPATH, '//*[@id="tux-6_input"]')
    full_name.send_keys(name)

    time.sleep(5)

    copyright_owner_name = driver.find_element(By.XPATH, '//*[@id="tux-7_input"]')
    copyright_owner_name.send_keys("maguybghosn")

    time.sleep(4)

    physical_address = driver.find_element(By.XPATH , '//*[@id="tux-8_input"]')
    physical_address.send_keys("Georgia")

    time.sleep(4)

    phone_nb = driver.find_element(By.XPATH, '//*[@id="tux-9_input"]')
    phone_nb.send_keys("+995577208104")

    time.sleep(5)

    option = driver.find_element(By.XPATH, '//*[@id="Video"]')
    option.click()

    time.sleep(2)

    copyright_src = driver.find_element(By.XPATH , '/html/body/div[3]/main/form/div[12]/div/div[2]/label/span')
    copyright_src.click()

    time.sleep(4)

    url_text1 = driver.find_element(By.XPATH , '//*[@id="tux-26_input"]') #content to report
    url_text1.send_keys("https://vt.tiktok.com/ZSFpqtJer/")

    time.sleep(4)

    url_text2 = driver.find_element(By.XPATH , '//*[@id="tux-27_textArea"]') #original
    url_text2.send_keys("https://vt.tiktok.com/ZSFpbrYa5/")

    time.sleep(4)

    desc_text = driver.find_element(By.XPATH , '//*[@id="tux-28_input"]')
    desc_text.send_keys("This account stole maguybghosn's video without her permission. Please consider this report.")

    time.sleep(5)

    upload_btn = driver.find_element(By.XPATH, '/html/body/div[3]/main/form/div[16]/div/label')
    upload_btn.click()
    time.sleep(1)
    pyautogui.write("ss.jpg")
    pyautogui.press('enter')
    assert WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".file"))).text == "ss.jpg"



    time.sleep(4)

    url_text3 = driver.find_element(By.XPATH , '//*[@id="tux-20_textArea"]')
    url_text3.send_keys("https://vt.tiktok.com/ZSFpbrYa5/") #original

    time.sleep(5)


    #prevent_btn = driver.find_element(By.XPATH, '/html/body/div[3]/main/form/div[16]/div/div/div/input')
    #prevent_btn.click()

    time.sleep(5)

    first_opt = driver.find_element(By.XPATH, '/html/body/div[3]/main/form/div[23]/div/div[1]/div')
    first_opt.click()

    time.sleep(5)

    second_opt = driver.find_element(By.XPATH, '/html/body/div[3]/main/form/div[23]/div/div[2]/div')
    second_opt.click()

    time.sleep(5)

    third_opt = driver.find_element(By.CSS_SELECTOR, 'div.css-ypesld:nth-child(3) > div:nth-child(2)')
    third_opt.click()

    time.sleep(5)

    signature = driver.find_element(By.ID , 'tux-25_input')
    signature.send_keys(name)

    time.sleep(5)

    send_btn = driver.find_element(By.XPATH, '/html/body/div[3]/main/form/button')
    send_btn.click()

    time.sleep(10)

    done_btn = driver.find_element(By.XPATH, '/html/body/div[3]/main/article/div/div/button')
    done_btn.click()

    time.sleep(5)
    
    driver.quit()

    driver.quit()

    time.sleep(20)





