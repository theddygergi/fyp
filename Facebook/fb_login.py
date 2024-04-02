from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
 
# Load workbook and select active sheet
workbook = openpyxl.load_workbook("email_password.xlsx")
worksheet = workbook.active
 
# Iterate over each row in the worksheet
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    # Extract email and password from each row
    email = row[0].value
    password = row[1].value
 
    # Initialize Firefox driver
    driver = webdriver.Firefox()
    driver.get("https://facebook.com")
 
    # Find email and password fields and enter credentials
    email_element = driver.find_element(By.ID, "email")
    password_element = driver.find_element(By.ID, "pass")
    email_element.send_keys(email)
    password_element.send_keys(password)
 
    # Click on the login button
    login_element = driver.find_element(By.NAME, "login")
    login_element.click()
 
    # Wait for some time to simulate a user interaction (you may adjust this)
    time.sleep(5)
 
    # Perform any actions needed after login
 
    # Close the browser
    driver.quit()